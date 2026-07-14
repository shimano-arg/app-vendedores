"""Cruce masivo Excel (formulario alta cliente) x Firestore client_applications
para cargar nombre de fantasia por CUIT.

Fuente: SHIMANO FISHING ARGENTINA (respuestas).xlsx
  col E = CUIT, col D = Nombre de fantasia

Target: coleccion client_applications - campo fantasia (si esta vacio o
igual al comercio, se pisa; si tiene algo distinto, se respeta salvo --force).

Uso:
  python bulk_import_fantasias_from_excel.py             # DRY RUN por default
  python bulk_import_fantasias_from_excel.py --apply     # ejecuta writes
  python bulk_import_fantasias_from_excel.py --apply --force  # pisa fantasias existentes
"""
import argparse
import json
import sys
from pathlib import Path
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

parser = argparse.ArgumentParser()
parser.add_argument('--apply', action='store_true', help='Ejecutar writes (default: dry run)')
parser.add_argument('--force', action='store_true', help='Pisar fantasias existentes distintas del Excel')
parser.add_argument('--xl', default=str(Path.home() / 'Downloads' / 'SHIMANO FISHING ARGENTINA (respuestas) (2).xlsx'))
args = parser.parse_args()

DRY = not args.apply
FORCE = args.force

print(f'MODO: {"DRY-RUN (solo mostrar)" if DRY else "APPLY (escribe a Firestore)"}')
print(f'FORCE: {FORCE}')
print(f'Excel: {args.xl}')

# =========================================================================
# 1. Leer Excel -> dict cuit -> fantasia
# =========================================================================
def norm_cuit(s):
    """Solo digitos. Excel a veces trae float con .0."""
    if s is None: return ''
    if isinstance(s, float): s = str(int(s))
    return ''.join(c for c in str(s) if c.isdigit())

def norm_fant(s):
    if s is None: return ''
    return str(s).strip()

wb = openpyxl.load_workbook(args.xl, data_only=True)
ws = wb[wb.sheetnames[0]]
excel_map = {}  # cuit -> {fantasia, titular, row_num}
dups = []
for row in range(3, ws.max_row + 1):
    cuit_raw = ws.cell(row=row, column=5).value
    fant_raw = ws.cell(row=row, column=4).value
    titular_raw = ws.cell(row=row, column=3).value
    cuit = norm_cuit(cuit_raw)
    fant = norm_fant(fant_raw)
    titular = norm_fant(titular_raw)
    if not cuit or not fant: continue
    if len(cuit) < 8: continue  # descartar CUITs malformados
    if cuit in excel_map:
        dups.append((row, cuit, fant, excel_map[cuit]['fantasia']))
        # Prefiero la ultima entrada (assumption: si el user cargo dos veces, la 2da es la buena)
    excel_map[cuit] = {'fantasia': fant, 'titular': titular, 'row_num': row}

print(f'\nExcel: {len(excel_map)} CUITs unicos con fantasia')
if dups:
    print(f'  {len(dups)} CUITs duplicados en el Excel (uso el ultimo, ver detalle):')
    for r, c, new, old in dups[:5]:
        print(f'    fila {r}: cuit={c} fantasia_nueva={new!r}  (previa={old!r})')

# =========================================================================
# 2. Cargar client_applications de Firestore
# =========================================================================
SA_KEY = Path.home() / 'Desktop' / 'sa-key.json'
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(json.loads(SA_KEY.read_text())))
db = firestore.client()

print(f'\nCargando client_applications de Firestore...')
apps = []
for d in db.collection('client_applications').stream():
    data = d.to_dict() or {}
    apps.append({'_id': d.id, '_ref': d.reference, **data})
print(f'  {len(apps)} docs total')

# =========================================================================
# 3. Match por CUIT + decidir accion
# =========================================================================
def app_cuit(a):
    """Prioridad: campo cuit -> extraer de cardCodeSap (formato C{cuit})."""
    c = norm_cuit(a.get('cuit', ''))
    if c and len(c) >= 8: return c
    # Fallback: cardCodeSap suele ser "C" + 11 digitos del CUIT
    cc = (a.get('cardCodeSap') or '').strip()
    if cc.startswith('C') or cc.startswith('c'):
        c2 = norm_cuit(cc[1:])
        if len(c2) >= 8: return c2
    return ''

stats = {
    'no_cuit_en_app': 0,
    'no_match_excel': 0,
    'already_ok': 0,       # ya tiene fantasia == excel
    'to_set_new': 0,       # fantasia vacia o == comercio, se puede setear
    'skip_manual': 0,      # tiene fantasia manual distinta, no pisar (sin --force)
    'to_overwrite': 0,     # tiene fantasia distinta y --force
}
actions = []  # (doc, current_fant, new_fant, reason)

for a in apps:
    cuit = app_cuit(a)
    if not cuit:
        stats['no_cuit_en_app'] += 1
        continue
    xl = excel_map.get(cuit)
    if not xl:
        stats['no_match_excel'] += 1
        continue
    new_fant = xl['fantasia']
    cur_fant = norm_fant(a.get('fantasia', ''))
    comercio = norm_fant(a.get('comercio', ''))
    # Consideramos "fantasia real cargada" si tiene algo != vacio y != comercio.
    has_real_fantasia = bool(cur_fant) and cur_fant.strip().lower() != comercio.strip().lower()
    same_as_excel = (cur_fant.strip().lower() == new_fant.strip().lower())

    if same_as_excel:
        stats['already_ok'] += 1
    elif not has_real_fantasia:
        stats['to_set_new'] += 1
        actions.append((a, cur_fant, new_fant, 'set (vacio o == comercio)'))
    elif FORCE:
        stats['to_overwrite'] += 1
        actions.append((a, cur_fant, new_fant, 'overwrite (--force)'))
    else:
        stats['skip_manual'] += 1

# =========================================================================
# 4. Mostrar stats + actions
# =========================================================================
print(f'\n{"="*70}')
print('STATS DEL CRUCE:')
print(f'{"="*70}')
for k, v in stats.items():
    print(f'  {k:22} = {v}')

print(f'\nAcciones a ejecutar: {len(actions)}')
if actions:
    print(f'\nSample (primeras 20):')
    for a, cur, new, reason in actions[:20]:
        print(f'  {a["_id"][:20]:22} cuit={app_cuit(a):12} comercio={norm_fant(a.get("comercio", ""))[:30]:32} -> fantasia={new!r} ({reason})')

# =========================================================================
# 5. Ejecutar writes (batch)
# =========================================================================
if not actions:
    print('\nNada para actualizar.')
    sys.exit(0)

if DRY:
    print(f'\n>>> DRY-RUN, no se escribio nada. Re-corre con --apply para ejecutar.')
    sys.exit(0)

print(f'\n>>> Aplicando {len(actions)} writes a Firestore...')
# Firestore batch soporta hasta 500 ops por batch
BATCH_LIMIT = 400
committed = 0
batch = db.batch()
count = 0
for a, cur, new, reason in actions:
    batch.update(a['_ref'], {
        'fantasia': new,
        'fantasiaSource': 'bulk_excel_2026-07-14',
        'fantasiaUpdatedAt': firestore.SERVER_TIMESTAMP,
    })
    count += 1
    if count >= BATCH_LIMIT:
        batch.commit()
        committed += count
        print(f'  commit batch: {committed} acumulados')
        batch = db.batch()
        count = 0
if count > 0:
    batch.commit()
    committed += count
print(f'\nOK: {committed} writes commit a client_applications.fantasia')
