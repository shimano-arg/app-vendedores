"""Corrige provincia + localidad en client_applications cruzando por CUIT
contra el Excel del formulario alta.

Detecta docs donde el sync SAP cargo provincia/localidad mal (ej:
TOMPY PESCA cargado como CHUBUT cuando es SALTA) y los pisa con el
valor correcto del Excel.

Uso:
  python bulk_fix_provincia_localidad_from_excel.py             # DRY RUN
  python bulk_fix_provincia_localidad_from_excel.py --apply     # ejecuta
"""
import argparse
import json
import sys
from pathlib import Path
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

parser = argparse.ArgumentParser()
parser.add_argument('--apply', action='store_true')
parser.add_argument('--xl', default=str(Path.home() / 'Downloads' / 'SHIMANO FISHING ARGENTINA (respuestas) (2).xlsx'))
args = parser.parse_args()

DRY = not args.apply
print(f'MODO: {"DRY-RUN" if DRY else "APPLY"}')

def _n_cuit(s):
    if s is None: return ''
    if isinstance(s, float): s = str(int(s))
    return ''.join(c for c in str(s) if c.isdigit())

def _n_str(s):
    return str(s or '').strip()

def _n_prov(s):
    # Normalizacion aggressive para comparar provincias (upper + sin acentos + trim)
    import unicodedata
    s = str(s or '').strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s

# Lista canonica de las 24 provincias/CABA argentinas + alias comunes de CABA.
# Solo aplicamos correcciones de provincia si el valor del Excel normalizado
# esta en esta lista. Descarta abreviaturas ("BS AS"), formatos raros
# ("CABA - CIUDAD AUTONOMA"), typos, etc. que corromperian data buena.
CANONICAL_PROVS = {
    'BUENOS AIRES',
    'CABA',  # alias
    'CIUDAD AUTONOMA DE BUENOS AIRES',
    'CAPITAL FEDERAL',  # alias
    'CATAMARCA',
    'CHACO',
    'CHUBUT',
    'CORDOBA',
    'CORRIENTES',
    'ENTRE RIOS',
    'FORMOSA',
    'JUJUY',
    'LA PAMPA',
    'LA RIOJA',
    'MENDOZA',
    'MISIONES',
    'NEUQUEN',
    'RIO NEGRO',
    'SALTA',
    'SAN JUAN',
    'SAN LUIS',
    'SANTA CRUZ',
    'SANTA FE',
    'SANTIAGO DEL ESTERO',
    'TIERRA DEL FUEGO',
    'TUCUMAN',
}
def is_canonical_prov(prov):
    return _n_prov(prov) in CANONICAL_PROVS

# ============= 1. Excel =============
wb = openpyxl.load_workbook(args.xl, data_only=True)
ws = wb[wb.sheetnames[0]]

# Provincia canonica: UPPERCASE del Excel. El Excel tiene variantes tipo
# "Salta", "Buenos Aires", "salta". Uppercase para el standard de la app.
excel_map = {}
for row in range(3, ws.max_row + 1):
    cuit = _n_cuit(ws.cell(row=row, column=5).value)
    if not cuit or len(cuit) < 8: continue
    prov_raw = _n_str(ws.cell(row=row, column=10).value)  # col J
    loc_raw = _n_str(ws.cell(row=row, column=9).value)    # col I
    prov = prov_raw.upper() if prov_raw else ''
    loc = loc_raw.upper() if loc_raw else ''
    if not prov and not loc: continue
    excel_map[cuit] = {'provincia': prov, 'localidad': loc, 'row': row}

print(f'Excel: {len(excel_map)} CUITs con provincia o localidad')

# ============= 2. Firestore =============
SA = Path.home() / 'Desktop' / 'sa-key.json'
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(json.loads(SA.read_text())))
db = firestore.client()
apps = []
for d in db.collection('client_applications').stream():
    apps.append({'_id': d.id, '_ref': d.reference, **(d.to_dict() or {})})
print(f'Firestore: {len(apps)} client_applications')

# ============= 3. Match + comparar =============
def app_cuit(a):
    c = _n_cuit(a.get('cuit', ''))
    if len(c) >= 8: return c
    cc = _n_str(a.get('cardCodeSap', ''))
    if cc and cc[0].upper() == 'C':
        c2 = _n_cuit(cc[1:])
        if len(c2) >= 8: return c2
    return ''

actions = []  # (a, {new_prov, new_loc, changes})
skipped_bad_prov = []  # docs con Excel prov no canonica - reportar
stats = {'no_cuit': 0, 'no_match_excel': 0, 'no_changes': 0, 'to_fix': 0,
         'skipped_prov_no_canonical': 0}

for a in apps:
    cuit = app_cuit(a)
    if not cuit:
        stats['no_cuit'] += 1
        continue
    xl = excel_map.get(cuit)
    if not xl:
        stats['no_match_excel'] += 1
        continue

    cur_prov = _n_str(a.get('provincia', ''))
    new_prov = xl['provincia']

    changes = {}
    # Provincia: solo tocar si (a) Excel tiene una PROVINCIA CANONICA y
    # (b) es distinta del valor actual normalizado. Esto descarta cambios
    # como "CIUDAD AUTONOMA..." -> "BS AS" (abreviatura no canonica que
    # corromperia el filtro por provincia de la app).
    if new_prov and _n_prov(cur_prov) != _n_prov(new_prov):
        if is_canonical_prov(new_prov):
            changes['provincia'] = {'from': cur_prov, 'to': new_prov}
        else:
            skipped_bad_prov.append({
                'cuit': cuit,
                'comercio': _n_str(a.get('comercio', ''))[:35],
                'cur': cur_prov,
                'excel': new_prov,
            })
            stats['skipped_prov_no_canonical'] += 1

    # v300+: NO tocar localidad automaticamente. El usuario del formulario
    # tipea muy variable ("Rosario Norte" vs "Rosario", "MAIPU" vs
    # "MAIPÚ", "LANSU OESTE" que es typo de LANUS, etc). Mejor dejar la
    # localidad al admin en Master Clientes.
    # Si mañana queremos, se agrega un flag --loc para volver a incluir.

    if not changes:
        stats['no_changes'] += 1
        continue

    stats['to_fix'] += 1
    actions.append({'app': a, 'cuit': cuit, 'changes': changes})

print()
print('=' * 70)
print('STATS')
print('=' * 70)
for k, v in stats.items():
    print(f'  {k:20} = {v}')

if actions:
    print(f'\nAcciones ({len(actions)}):')
    for i, act in enumerate(actions, 1):
        a = act['app']
        c = act['changes']
        chg_parts = []
        for field, cv in c.items():
            chg_parts.append(f'{field}: {cv["from"]!r} -> {cv["to"]!r}')
        print(f'  {i:>3}. {act["cuit"]}  {_n_str(a.get("comercio", ""))[:30]:32}  {"  |  ".join(chg_parts)}')

if skipped_bad_prov:
    print(f'\nSkipped (Excel prov NO canonica, no aplico): {len(skipped_bad_prov)}')
    for s in skipped_bad_prov[:10]:
        print(f'  {s["cuit"]}  {s["comercio"]:35}  cur={s["cur"]!r:35}  excel={s["excel"]!r}')

if DRY:
    print('\n>>> DRY-RUN, nada se escribio. Re-corre con --apply.')
    sys.exit(0)

if not actions:
    print('\nNada para actualizar.')
    sys.exit(0)

print(f'\n>>> Aplicando {len(actions)} writes a Firestore...')
BATCH = 400
batch = db.batch()
count = 0
committed = 0
for act in actions:
    upd = {'updatedAt': firestore.SERVER_TIMESTAMP}
    if 'provincia' in act['changes']:
        upd['provincia'] = act['changes']['provincia']['to']
    if 'localidad' in act['changes']:
        upd['localidad'] = act['changes']['localidad']['to']
        upd['localidadFinal'] = act['changes']['localidad']['to']
    # Audit
    upd['provinciaLocSource'] = 'bulk_excel_2026-07-14'
    upd['provinciaLocUpdatedAt'] = firestore.SERVER_TIMESTAMP
    batch.update(act['app']['_ref'], upd)
    count += 1
    if count >= BATCH:
        batch.commit()
        committed += count
        print(f'  commit: {committed} acumulados')
        batch = db.batch()
        count = 0
if count > 0:
    batch.commit()
    committed += count
print(f'\nOK: {committed} writes')
