"""Lee el Excel SHIMANO FISHING ARGENTINA para inspeccionar estructura
antes de armar el matching CUIT -> fantasia."""
import openpyxl
from pathlib import Path

XL = Path.home() / 'Downloads' / 'SHIMANO FISHING ARGENTINA (respuestas) (2).xlsx'
print(f'Leyendo: {XL}')
print(f'Existe: {XL.exists()}')

wb = openpyxl.load_workbook(XL, data_only=True)
print(f'Hojas: {wb.sheetnames}')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print()
    print('=' * 70)
    print(f'HOJA: {sheet_name}  ({ws.max_row} filas x {ws.max_column} cols)')
    print('=' * 70)
    # Headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
    print(f'\nHeaders:')
    for i, h in enumerate(headers, 1):
        col_letter = chr(64 + i) if i <= 26 else '?'
        print(f'  {col_letter} col{i:>2}  {h}')

    # Primeras 5 filas de data
    print(f'\nPrimeras 5 filas de data:')
    for row in range(2, min(7, ws.max_row + 1)):
        print(f'\n  --- fila {row} ---')
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            header = headers[col-1] if col-1 < len(headers) else f'col{col}'
            val_str = str(val)[:60] if val is not None else '(vacio)'
            print(f'    {str(header)[:30]:<32} = {val_str}')

    # Contar filas con CUIT poblado (columna E = col 5 segun screenshot)
    print(f'\nStats:')
    cuit_col_idx = None
    fant_col_idx = None
    for i, h in enumerate(headers):
        if h and 'cuit' in str(h).lower():
            cuit_col_idx = i + 1
        if h and 'fantas' in str(h).lower():
            fant_col_idx = i + 1
    print(f'  CUIT en col {cuit_col_idx}')
    print(f'  Fantasia en col {fant_col_idx}')
    if cuit_col_idx and fant_col_idx:
        con_cuit = 0
        con_fant = 0
        con_ambos = 0
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=cuit_col_idx).value
            f = ws.cell(row=row, column=fant_col_idx).value
            if c: con_cuit += 1
            if f: con_fant += 1
            if c and f: con_ambos += 1
        print(f'  Filas con CUIT: {con_cuit}')
        print(f'  Filas con Fantasia: {con_fant}')
        print(f'  Filas con AMBOS: {con_ambos} <-- candidatos al match')
