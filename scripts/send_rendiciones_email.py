"""
Send Rendiciones Aprobadas - cron Lun/Mie 9 AM AR
==================================================

Lee Firestore las rendiciones que estan APROBADAS y aun no notificadas
(notifiedAt == null), arma un Excel y lo manda por mail al admin.
Despues marca cada rendicion como notifiedAt=now para no duplicar.

Variables de entorno (cargadas por GitHub Actions):
    FIREBASE_SERVICE_ACCOUNT  JSON del service account (string).
    GMAIL_APP_PASSWORD        16 chars del App Password.
    MAIL_FROM                 bot.shimano.pesca@gmail.com
    MAIL_TO                   mariano.erbino@shimano.com.ar

Logica:
  1. Conectarse a Firestore con el service account.
  2. Query: collection 'rendiciones' where status='approved'
     and (notifiedAt does not exist OR notifiedAt is null).
  3. Subir TODAS las fotos a Firebase Storage y cachear URLs por rendicion id.
  4. Generar Excel con 3 hojas:
       - "Gastos": agrupado por (ownerEmail, tipoGasto). Una fila por dupla
         con importeTotal + cantRendiciones + fotosUrls (separadas por ';') +
         rendicionesIds. Tabla "TablaGastos" que lee Power Automate.
       - "Detalle": una fila por gasto individual, formato viejo (audit only).
       - "Solicitudes": sin agrupar, una fila por anticipo. Tabla
         "TablaSolicitudes".
  5. Si hay >= 1 rendicion para notificar, mandar mail con el xlsx
     adjunto. Si no hay nada, salir sin mandar mail (evita spam).
  6. Marcar las notificadas en batch.

Cambio v2 (2026-06-30, Fernando):
  Antes Power Automate creaba 1 item SharePoint POR rendicion. Si Gonzalo
  tenia 3 facturas A salian 3 items separados, complicando la rendicion.
  Ahora el script agrupa: 1 fila = 1 item SharePoint con la suma y todas
  las fotos como adjuntos. Power Automate hace split de fotosUrls y attach
  uno por uno.

Cambio v3 (2026-07-29, Mariano):
  Antes el flow adjuntaba el Excel MAESTRO completo a cada item SharePoint
  (todos los vendedores en un solo xlsx). Al abrir la fila de Federico el
  Excel adjunto tenia las 5 rendiciones de todos, no solo las de Federico.
  Ahora el script genera N mini-Excels (uno por dupla vendedor+tipoGasto)
  y los sube a Firebase Storage. La URL publica se incluye en la nueva
  columna "Excel Dupla URL" de TablaGastos. El flow hace HTTP GET a esa
  URL + SharePoint Add attachment para poner el xlsx filtrado en el
  item correspondiente. El Excel maestro sigue yendo al mail como
  resumen para Mariano/Fernando.
"""

from __future__ import annotations

import base64
import io
import json
import os
import re
import smtplib
import sys
from collections import defaultdict
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path

import firebase_admin
from firebase_admin import credentials, firestore, storage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

MAIL_FROM = os.environ.get("MAIL_FROM", "bot.shimano.pesca@gmail.com")
MAIL_TO = os.environ.get("MAIL_TO", "mariano.erbino@shimano.com.ar")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
FB_SA_JSON = os.environ.get("FIREBASE_SERVICE_ACCOUNT", "")

def _envbool(name: str) -> bool:
    return (os.environ.get(name, "") or "").strip().lower() in ("1", "true", "yes", "on")

FORCE_SEND = _envbool("FORCE_SEND")  # ignora filtro notifiedAt
SKIP_MARK = _envbool("SKIP_MARK")    # NO marca como notificadas (testing)
REPLAY_IDS = [x.strip() for x in (os.environ.get("REPLAY_IDS", "") or "").split(",") if x.strip()]


def die(msg: str) -> None:
    print(f"::error::{msg}", file=sys.stderr)
    sys.exit(1)


def init_firestore():
    if not FB_SA_JSON:
        die("FIREBASE_SERVICE_ACCOUNT vacio. Setear el secret en GitHub.")
    try:
        sa_dict = json.loads(FB_SA_JSON)
    except json.JSONDecodeError as e:
        die(f"FIREBASE_SERVICE_ACCOUNT no es JSON valido: {e}")
    project_id = sa_dict.get("project_id", "")
    # Default storage bucket: <project-id>.firebasestorage.app (formato nuevo,
    # post 2024). Antes era <project-id>.appspot.com. Si el bucket esta
    # deshabilitado/no inicializado, el upload falla con 404 'bucket does
    # not exist' y caemos al fallback de campo vacio (URL omitida).
    # Override via env var STORAGE_BUCKET si Firebase usa el nombre legacy.
    default_bucket = f"{project_id}.firebasestorage.app" if project_id else None
    storage_bucket = os.environ.get("STORAGE_BUCKET") or default_bucket
    cred = credentials.Certificate(sa_dict)
    options = {"storageBucket": storage_bucket} if storage_bucket else None
    firebase_admin.initialize_app(cred, options)
    return firestore.client()


def upload_foto_to_storage(rendicion_id: str, foto_dataurl: str):
    """Sube la foto a Firebase Storage en un path predecible y devuelve la URL
    publica permanente. None si no se pudo (storage deshabilitado o foto
    corrupta)."""
    if not foto_dataurl or not isinstance(foto_dataurl, str):
        return None
    try:
        m = re.match(r"^data:image/([\w+]+);base64,(.+)$", foto_dataurl, re.IGNORECASE)
        if m:
            ext = m.group(1).lower()
            b64_data = m.group(2)
        else:
            # No es dataURL: asumir base64 puro + jpeg.
            ext = "jpeg"
            b64_data = foto_dataurl
        if ext == "jpg":
            ext = "jpeg"
        raw = base64.b64decode(b64_data)
        bucket = storage.bucket()
        blob = bucket.blob(f"rendiciones-tickets/{rendicion_id}.{ext}")
        blob.upload_from_string(raw, content_type=f"image/{ext}")
        blob.make_public()
        return blob.public_url
    except Exception as e:
        print(f"[storage] no pude subir foto {rendicion_id}: {e}", file=sys.stderr)
        return None


def _slug(text: str) -> str:
    """Slug simple para path de Storage: quita chars no [a-z0-9-_.] y trunca."""
    s = re.sub(r"[^\w.@-]+", "_", (text or "").strip().lower())
    return s[:80] or "x"


def upload_xlsx_to_storage(email: str, tipo: str, xlsx_bytes: bytes, day_str: str):
    """Sube el Excel filtrado (una dupla vendedor+tipoGasto) a Firebase Storage
    y devuelve la URL publica permanente. Path predecible por dia + dupla —
    si el cron se corre 2 veces el mismo dia, el segundo run pisa el archivo
    del primero (no acumula). None si no se pudo subir."""
    if not xlsx_bytes:
        return None
    try:
        key = f"{_slug(email)}__{_slug(tipo)}"
        blob = storage.bucket().blob(f"rendiciones-excels/{day_str}/{key}.xlsx")
        blob.upload_from_string(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        blob.make_public()
        return blob.public_url
    except Exception as e:
        print(f"[storage] no pude subir xlsx {email}/{tipo}: {e}", file=sys.stderr)
        return None


def _build_dupla_xlsx(items: list, foto_url_by_id: dict, email: str, tipo: str) -> bytes:
    """Arma un Excel MINI con solo las rendiciones de una dupla (vendedor, tipoGasto).
    2 hojas: "Resumen" (una fila con totales) + "Detalle" (una fila por gasto).
    Estilo consistente con el Excel maestro pero solo con las filas de la dupla —
    asi Fernando abre desde SharePoint la fila de Federico Factura A y ve solo
    las 2 rendiciones de esa dupla, no las 5 de todos los vendedores."""
    wb = Workbook()

    # === HOJA 1: RESUMEN (misma info que la fila de TablaGastos maestra) ===
    ws_r = wb.active
    ws_r.title = "Resumen"
    hdr_r = ["Vendedor (email)", "Tipo gasto", "Cant Rendiciones",
             "Importe Total", "Importe USD Total", "Moneda",
             "Periodo Desde", "Periodo Hasta"]
    ws_r.append(hdr_r)
    for cell in ws_r[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")
    cant = len(items)
    importe_total = sum(float(r.get("importe") or 0) for r in items)
    importe_usd_total = sum(float(r.get("importeUsd") or 0) for r in items if r.get("importeUsd"))
    monedas = {(r.get("moneda") or "").strip() for r in items if r.get("moneda")}
    moneda_lbl = next(iter(monedas)) if len(monedas) == 1 else ("MIXTO" if monedas else "")
    fechas = sorted(filter(None, (fmt_ts(r.get("createdAt")) for r in items)))
    desde = fechas[0] if fechas else ""
    hasta = fechas[-1] if fechas else ""
    ws_r.append([
        email, tipo, cant,
        round(importe_total, 2),
        round(importe_usd_total, 2) if importe_usd_total else "",
        moneda_lbl, desde, hasta,
    ])
    for i, w in enumerate([28, 26, 14, 16, 16, 10, 18, 18]):
        ws_r.column_dimensions[get_column_letter(i + 1)].width = w

    # === HOJA 2: DETALLE (una fila por rendicion de esta dupla) ===
    ws_d = wb.create_sheet("Detalle")
    hdr_d = [
        "ID", "Fecha carga", "Vendedor (email)", "N° Ticket", "Descripcion",
        "Modo pago", "Tipo gasto", "Division gasto", "Moneda", "Importe",
        "Importe USD", "Observaciones", "Aprobado por", "Fecha aprobacion",
        "Ticket",
    ]
    ws_d.append(hdr_d)
    for cell in ws_d[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")
    foto_col_idx_d = len(hdr_d)
    link_font = Font(color="0563C1", underline="single", bold=True)
    for r in items:
        ws_d.append([
            r.get("_id", ""),
            fmt_ts(r.get("createdAt")),
            r.get("ownerEmail") or r.get("createdByEmail") or "",
            r.get("numeroTicket") or "",
            r.get("descripcion") or "",
            r.get("modoPago") or "",
            r.get("tipoGasto") or "",
            r.get("divisionGasto") or "",
            r.get("moneda") or "",
            float(r.get("importe") or 0),
            float(r.get("importeUsd") or 0) if r.get("importeUsd") else "",
            r.get("observaciones") or "",
            r.get("approvedByEmail") or "",
            fmt_ts(r.get("approvedAt")),
            "",  # placeholder
        ])
        cell = ws_d.cell(row=ws_d.max_row, column=foto_col_idx_d)
        rid = r.get("_id", "")
        url = foto_url_by_id.get(rid)
        if url:
            cell.value = "📷 Ver ticket"
            cell.hyperlink = url
            cell.font = link_font
            cell.alignment = Alignment(horizontal="center")
        elif r.get("fotoTicket") or r.get("adjunto"):
            cell.value = "(error al subir)"
        else:
            cell.value = "(sin foto)"
    for i, w in enumerate([22, 16, 28, 14, 26, 14, 22, 16, 12, 12, 12, 36, 26, 16, 14]):
        ws_d.column_dimensions[get_column_letter(i + 1)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fetch_pending_approved(db):
    """Query rendiciones aprobadas. Si FORCE_SEND, traemos todas las
    aprobadas (incluso las ya notificadas) - util para testear el formato
    del Excel sin tener que esperar nuevas rendiciones. Si REPLAY_IDS
    tiene contenido, ignoramos status/notifiedAt y traemos solo esos IDs
    (replay quirurgico de un batch puntual, ej. cuando un cron corrio con
    un bug y hay que reenviar solo ese subconjunto sin duplicar historico
    en SharePoint via Power Automate)."""
    if REPLAY_IDS:
        out = []
        coll = db.collection("rendiciones")
        for rid in REPLAY_IDS:
            snap = coll.document(rid).get()
            if not snap.exists:
                print(f"[fetch] WARN: REPLAY_ID {rid} no existe", file=sys.stderr)
                continue
            data = snap.to_dict() or {}
            data["_id"] = snap.id
            data["_ref"] = snap.reference
            out.append(data)
        return out
    docs = db.collection("rendiciones").where("status", "==", "approved").stream()
    out = []
    for d in docs:
        data = d.to_dict() or {}
        if not FORCE_SEND:
            # Filtrar las que YA fueron notificadas (notifiedAt no null).
            notified_at = data.get("notifiedAt")
            if notified_at:
                continue
        data["_id"] = d.id
        data["_ref"] = d.reference
        out.append(data)
    return out


def fmt_ts(ts):
    """Convierte un Firestore Timestamp / datetime / string a 'YYYY-MM-DD HH:MM'."""
    if not ts:
        return ""
    try:
        if hasattr(ts, "to_datetime"):
            dt = ts.to_datetime()
        elif hasattr(ts, "isoformat"):
            dt = ts
        else:
            return str(ts)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(ts)


def build_excel(rendiciones):
    """Arma el Workbook con 3 hojas: Gastos (agrupado), Detalle, Solicitudes.

    Layout v2 (a pedido de Fernando, 2026-06-30):
      Hoja "Gastos": una fila por (ownerEmail, tipoGasto). Suma importes y
        concatena fotosUrls + rendicionesIds. Power Automate lee TablaGastos
        y crea UN item SharePoint por fila + adjunta todas las fotos.
      Hoja "Detalle": una fila por gasto individual (vista ungroupeada).
        Solo para auditoria del que abre el Excel - NO se mapea a SharePoint.
      Hoja "Solicitudes": una fila por solicitud (sin agrupar - cada anticipo
        es independiente).
    """
    wb = Workbook()
    gastos = [r for r in rendiciones if r.get("tipo") == "gasto"]
    sols = [r for r in rendiciones if r.get("tipo") == "solicitud"]

    # PASO 1: obtener URL publica de la foto de cada rendicion.
    # 2 casos segun cuando se creo la rendicion:
    #   a) v308+: la app subio la foto a Firebase Storage al crear la rendicion
    #      y guardo la URL en `fotoTicketUrl`. Reusamos esa URL directo (skip
    #      re-upload). Este es el caso mayoritario para rendiciones nuevas.
    #   b) pre-v308: la foto vive como base64 dataURL en `fotoTicket` (o
    #      `adjunto` legacy). Subimos a Storage acá para obtener la URL.
    # Sin este dispatch, las rendiciones post-v308 aparecen como "sin foto" en
    # el Excel porque el field `fotoTicket` esta vacio - toda la data esta en
    # `fotoTicketUrl`. Bug observado 2026-07-27: rendicion de diego.valsi con
    # foto valida en la app pero flow de Power Automate fallo por Fotos URLs
    # vacio (fix del flow ya aplicado, este fix cierra el otro lado).
    foto_url_by_id = {}
    for r in gastos:
        rid = r.get("_id", "rend")
        existing_url = r.get("fotoTicketUrl")
        if existing_url:
            foto_url_by_id[rid] = existing_url
            continue
        foto_src = r.get("fotoTicket") or r.get("adjunto") or ""
        if not foto_src:
            continue
        url = upload_foto_to_storage(rid, foto_src)
        if url:
            foto_url_by_id[rid] = url

    # PASO 2: agrupar por (ownerEmail, tipoGasto). Estos son los 3 valores
    # tipicos: "Factura A", "Gastos con comprobante", "Gastos sin comprobante".
    # Si una persona tiene 3 rendiciones de Factura A -> 1 fila con suma.
    groups = defaultdict(list)
    for r in gastos:
        email = (r.get("ownerEmail") or r.get("createdByEmail") or "").strip()
        tipo = (r.get("tipoGasto") or "").strip() or "(sin tipo)"
        groups[(email, tipo)].append(r)

    # PASO 2.5 (v3, 2026-07-29): por cada dupla, generar un mini-Excel con
    # solo sus rendiciones + subirlo a Firebase Storage. La URL publica se
    # incluye en una columna nueva "Excel Dupla URL" que Power Automate
    # lee para adjuntar al item SharePoint (HTTP GET + Add attachment).
    # Antes: el flow adjuntaba el Excel MAESTRO completo a cada item —
    # la fila de Federico tenia adjunto el Excel con las 5 rendiciones
    # de todos los vendedores. Ahora la fila de Federico Factura A tiene
    # adjunto un Excel de 8 KB con solo sus 2 rendiciones de Factura A.
    day_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dupla_xlsx_url = {}
    for (email, tipo), items in groups.items():
        try:
            xlsx_dupla = _build_dupla_xlsx(items, foto_url_by_id, email, tipo)
            url = upload_xlsx_to_storage(email, tipo, xlsx_dupla, day_str)
            if url:
                dupla_xlsx_url[(email, tipo)] = url
        except Exception as e:
            print(f"[storage] error generando xlsx dupla {email}/{tipo}: {e}", file=sys.stderr)

    # === HOJA 1: GASTOS AGRUPADO (la que lee Power Automate) ===
    ws_g = wb.active
    ws_g.title = "Gastos"
    hdr_g = [
        "Vendedor (email)", "Tipo gasto", "Cant Rendiciones",
        "Importe Total", "Importe USD Total", "Moneda",
        "Periodo Desde", "Periodo Hasta",
        "Rendiciones IDs", "Fotos URLs",
        "Excel Dupla URL",  # v3 (2026-07-29): URL del xlsx filtrado por dupla
    ]
    ws_g.append(hdr_g)
    for cell in ws_g[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")

    for (email, tipo), items in sorted(groups.items(), key=lambda kv: (kv[0][0].lower(), kv[0][1].lower())):
        cant = len(items)
        importe_total = sum(float(r.get("importe") or 0) for r in items)
        importe_usd_total = sum(float(r.get("importeUsd") or 0) for r in items if r.get("importeUsd"))
        monedas = {(r.get("moneda") or "").strip() for r in items if r.get("moneda")}
        moneda_lbl = next(iter(monedas)) if len(monedas) == 1 else ("MIXTO" if monedas else "")
        fechas = sorted(filter(None, (fmt_ts(r.get("createdAt")) for r in items)))
        desde = fechas[0] if fechas else ""
        hasta = fechas[-1] if fechas else ""
        ids = ";".join(r.get("_id", "") for r in items)
        # Concatenar URLs solo de los gastos cuya foto SI se subio bien.
        # Si una foto fallo upload, se omite (no rompe la fila).
        urls = ";".join(foto_url_by_id[r["_id"]] for r in items if r.get("_id") in foto_url_by_id)
        excel_url = dupla_xlsx_url.get((email, tipo), "")
        ws_g.append([
            email, tipo, cant,
            round(importe_total, 2),
            round(importe_usd_total, 2) if importe_usd_total else "",
            moneda_lbl,
            desde, hasta,
            ids, urls,
            excel_url,
        ])
    # Anchos: Fotos URLs + Excel Dupla URL anchas porque son URLs largas.
    widths_g = [28, 26, 14, 16, 16, 10, 18, 18, 50, 70, 70]
    for i, w in enumerate(widths_g):
        ws_g.column_dimensions[get_column_letter(i + 1)].width = w
    # Power Automate lee TablaGastos. El displayName se mantiene para no
    # romper el flow existente - solo cambian las columnas dentro.
    # Fix 2026-08-05: mismo patron que TablaSolicitudes - si la semana no
    # tuvo gastos aprobados, agregamos 1 fila placeholder para que la tabla
    # exista siempre y el flow Power Automate no falle.
    if ws_g.max_row < 2:
        ws_g.append([""] * len(hdr_g))
    last_col_letter = get_column_letter(len(hdr_g))
    table_ref = f"A1:{last_col_letter}{ws_g.max_row}"
    tab_g = Table(displayName="TablaGastos", ref=table_ref)
    tab_g.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_g.add_table(tab_g)

    # === HOJA 2: DETALLE (auditoria, una fila por gasto) ===
    # No se mapea a SharePoint - es solo para que Fernando o vos puedan abrir
    # el Excel y ver linea por linea si necesitan investigar un monto.
    ws_d = wb.create_sheet("Detalle")
    hdr_d = [
        "ID", "Fecha carga", "Vendedor (email)", "N° Ticket", "Descripcion",
        "Modo pago", "Tipo gasto", "Division gasto", "Moneda", "Importe",
        "Importe USD", "Observaciones", "Aprobado por", "Fecha aprobacion",
        "Ticket",
    ]
    ws_d.append(hdr_d)
    for cell in ws_d[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")
    foto_col_idx_d = len(hdr_d)
    link_font = Font(color="0563C1", underline="single", bold=True)
    for r in gastos:
        ws_d.append([
            r.get("_id", ""),
            fmt_ts(r.get("createdAt")),
            r.get("ownerEmail") or r.get("createdByEmail") or "",
            r.get("numeroTicket") or "",
            r.get("descripcion") or "",
            r.get("modoPago") or "",
            r.get("tipoGasto") or "",
            r.get("divisionGasto") or "",
            r.get("moneda") or "",
            float(r.get("importe") or 0),
            float(r.get("importeUsd") or 0) if r.get("importeUsd") else "",
            r.get("observaciones") or "",
            r.get("approvedByEmail") or "",
            fmt_ts(r.get("approvedAt")),
            "",  # placeholder
        ])
        cell = ws_d.cell(row=ws_d.max_row, column=foto_col_idx_d)
        rid = r.get("_id", "")
        url = foto_url_by_id.get(rid)
        if url:
            cell.value = "📷 Ver ticket"
            cell.hyperlink = url
            cell.font = link_font
            cell.alignment = Alignment(horizontal="center")
        elif r.get("fotoTicket") or r.get("adjunto"):
            cell.value = "(error al subir)"
        else:
            cell.value = "(sin foto)"
    widths_d = [22, 16, 28, 14, 26, 14, 22, 16, 12, 12, 12, 36, 26, 16, 14]
    for i, w in enumerate(widths_d):
        ws_d.column_dimensions[get_column_letter(i + 1)].width = w

    # === HOJA 3: SOLICITUDES (sin agrupar, una fila por anticipo) ===
    ws_s = wb.create_sheet("Solicitudes")
    hdr_s = [
        "ID", "Fecha carga", "Vendedor (email)", "Solicitado por",
        "Tipo operacion", "Motivo", "Moneda", "Importe",
        "Estado", "Observaciones", "Aprobado por", "Fecha aprobacion",
    ]
    ws_s.append(hdr_s)
    for cell in ws_s[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")
    for r in sols:
        ws_s.append([
            r.get("_id", ""),
            fmt_ts(r.get("createdAt")),
            r.get("ownerEmail") or r.get("createdByEmail") or "",
            r.get("solicitadoPor") or "",
            r.get("tipoOperacion") or "",
            r.get("motivo") or "",
            r.get("moneda") or "",
            float(r.get("importe") or 0),
            r.get("estado") or "",
            r.get("observaciones") or "",
            r.get("approvedByEmail") or "",
            fmt_ts(r.get("approvedAt")),
        ])
    widths_s = [22, 16, 28, 22, 22, 30, 12, 12, 12, 36, 26, 16]
    for i, w in enumerate(widths_s):
        ws_s.column_dimensions[chr(65 + i)].width = w
    # Tabla con nombre tambien en Solicitudes (mismo motivo: Power Automate).
    # v760 (2026-09-01): FIX del "fix" 2026-08-05. El fix original agregaba una
    # fila placeholder vacia asumiendo que Power Automate iteraba 0 filas
    # utiles. FALSO — PA itera esa fila con todos los campos vacios y trata de
    # crear un item basura en SharePoint, que falla con BadRequest
    # ("No se encontro el usuario especificado" — porque Solicitado por Claims
    # es un People field y "" no es un usuario valido). Sintoma: replay con
    # solo gastos siempre fallaba.
    # Nuevo approach: si no hay solicitudes reales, NO crear la tabla. El flow
    # Power Automate tiene `Configure run after` tolerante en `List rows
    # Solicitudes` (`is successful + has failed + is skipped`) desde el fix
    # original — eso maneja el `NotFound` gracefully.
    if ws_s.max_row >= 2:
        last_col_letter_s = get_column_letter(len(hdr_s))
        table_ref_s = f"A1:{last_col_letter_s}{ws_s.max_row}"
        tab_s = Table(displayName="TablaSolicitudes", ref=table_ref_s)
        tab_s.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws_s.add_table(tab_s)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_email(xlsx_bytes: bytes, count: int) -> None:
    if not GMAIL_APP_PASSWORD:
        die("GMAIL_APP_PASSWORD vacio. Setear el secret en GitHub.")
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    subject = f"[Shimano App] Rendiciones aprobadas - {today_str} ({count} pendientes)"
    # v609 (2026-08-24): body simplificado a pedido de Mariano. Antes tenia
    # explicacion de las hojas del Excel; ahora solo el mensaje corto.
    body = (
        "Buenas Marian/Die,\n\n"
        "aca estan las rendiciones aprobadas por Pablo y que necesitan validacion.\n\n"
        "-- Shimano App Vendedores"
    )
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = MAIL_FROM
    msg["To"] = MAIL_TO
    msg.set_content(body)
    fname = f"Rendiciones_Aprobadas_{today_str}.xlsx"
    msg.add_attachment(
        xlsx_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
    )
    print(f"[mail] Enviando a {MAIL_TO} desde {MAIL_FROM}...")
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
        srv.login(MAIL_FROM, GMAIL_APP_PASSWORD)
        srv.send_message(msg)
    print(f"[mail] OK - asunto: {subject}")


def mark_as_notified(rendiciones) -> None:
    """Marca en Firestore cada rendicion como notifiedAt=now."""
    now = datetime.now(timezone.utc)
    for r in rendiciones:
        try:
            r["_ref"].update({
                "notifiedAt": now,
                "notifiedTo": MAIL_TO,
            })
        except Exception as e:
            print(f"[mark] error en {r.get('_id')}: {e}", file=sys.stderr)


def main() -> int:
    db = init_firestore()
    if REPLAY_IDS:
        mode_lbl = f"REPLAY_IDS ({len(REPLAY_IDS)} IDs puntuales)"
    elif FORCE_SEND:
        mode_lbl = "FORCE_SEND (todas)"
    else:
        mode_lbl = "solo no notificadas"
    print(f"[fetch] Modo: {mode_lbl}")
    rendiciones = fetch_pending_approved(db)
    print(f"[fetch] Encontradas: {len(rendiciones)}")
    if not rendiciones:
        print("[exit] Sin rendiciones para notificar. No envio mail.")
        return 0
    xlsx_bytes = build_excel(rendiciones)
    send_email(xlsx_bytes, len(rendiciones))
    if SKIP_MARK:
        print(f"[done] SKIP_MARK activo - {len(rendiciones)} rendiciones NO se marcaron como notificadas (testing).")
    else:
        mark_as_notified(rendiciones)
        print(f"[done] {len(rendiciones)} rendiciones marcadas como notificadas.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
